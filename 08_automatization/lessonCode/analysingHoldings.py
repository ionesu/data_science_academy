import openpyxl as excl
from lessonCode import writeFile, execution as e
import os

IDENTATION = "___"
exclusions = ['оборот', '62', '62.01', '62.02', 'начальное сальдо', 'конечное сальдо']
holdingsCos = ['бета ооо', 'альфа ооо', 'гамма ооо']

inputFolder = 'inputData/'

banksSummary = {}


# Есть слово банк, есть запятая, есть пробел. Первый элемент есть счет, который нам интересен
def processBank(contents):
    rowContents = contents.split(' ')
    account = rowContents[0]
    account = account.replace(',', '')
    return account


def runAnalysis():
    for file in os.listdir(inputFolder):
        book = excl.load_workbook(inputFolder + file)
        page = book['TDSheet']
        colA = page['A']
        companyName = colA[0].value

        entries = {}
        paymentStartIndex = 0

        banks = defineTheRange(entries, colA, companyName, paymentStartIndex)
        # Первая клетка – компания
        # какие строчки рассмартривать
        # Для выбранных строчек из шага 1:
        for bank in banks:

            monthlyIncome = []
            monthSummary = {}
            firstRow = banks[bank][0]
            lastRow = banks[bank][1]
            # sum = 0
            # sumHoldingCos = 0

            for row in page.iter_rows(min_row=firstRow, max_row=lastRow):
                # 1.Месяц определяется словом ‘оборот’
                if row[0].value and "оборот" in row[0].value.lower():
                    monthlyIncome.append(monthSummary)
                    # 2.Каждый месяц
                    monthSummary = {}

                    # 3.Если не мета-данные, смотрим на дебит
                if row[2].value.lower() not in exclusions:
                    # 4.Делаем общую сумму
                    monthSummary[companyName] = monthSummary.get(companyName, 0) + int(row[3].value)

                    # 5.Делаем сумму по холдинговым компаниям
                    if row[2].value.lower() in holdingsCos:
                        holdingsCoName = row[2].value.lower()
                        monthSummary[holdingsCoName] = monthSummary.get(holdingsCoName, 0) + int(row[3].value)

            banksSummary[bank] = monthlyIncome

        print(banksSummary)
        writeFile.writeSheet(excl.Workbook(), 'Результат', banksSummary)
    # cохранить результат


# Определяем какие строчки рассматривать на основе колнки А
def defineTheRange(banks, colA, companyName, paymentStartIndex):
    for rowNumber, cell in enumerate(colA):
        if (cell.value):
            # если есть слово 'банк'
            if 'банк ' in cell.value.lower():
                account = processBank(cell.value)
                accountWCoName = account + IDENTATION + companyName
                # print(accountWCoName)

            # Мы начинаем смотреть на поступления начиная со строки в которой есть ‘поступлен’ и ‘продаж’
            if "поступлен" in cell.value.lower() and "продаж" in cell.value.lower():
                paymentStartIndex = rowNumber
                # print(paymentStartIndex)

            elif paymentStartIndex > 0 and 'оборот' not in cell.value.lower():
                banks[accountWCoName] = [paymentStartIndex, rowNumber]
                # print(banks)
                paymentStartIndex = 0

    return banks


runAnalysis()
