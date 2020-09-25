import openpyxl.styles as cStyle
from lessonCode import execution as e


# Проверяем, что мы не работаем с компание по который мы делаем репорт
def isNotReportingCo(companyName, selfCompany):
    companyName = companyName.replace('"', '')  # убираем кавычки
    selfCompany = selfCompany.replace('"', '')  # убираем кавычки
    companyNameSorted = ''.join(sorted(companyName.lower()))  # сортируем по порядку, чтобы порядок не имел значение
    selfCompanySorted = ''.join(sorted(selfCompany.lower()))  # сортируем по порядку, чтобы порядок не имел значение
    if companyNameSorted not in selfCompanySorted:
        return True
    return False


def headerRowFormat(cell):
    cell.fill = cStyle.PatternFill(fgColor="80a9ed", fill_type="solid")
    cell.font = cStyle.Font(color="FFFFFF", bold=True, size=8)


# выписываем резульаты по месяцам
def writeMonthResults(dataRow, dataColumn, banksSummary, bank, resultSheet, param):
    for monthN in e.monInt:
        cell = resultSheet.cell(row=dataRow, column=dataColumn)
        totalIncome = banksSummary[bank][monthN].get(param, 0)  # если значения нет, выдаем ноль
        cell.value = '{:,}'.format(totalIncome)  # форматируем число
        dataColumn += 1  # идем на следующую колонку


# проверям что компания производила транзакцию по этому счету
def existsInReport(companyName, monthlyResults):
    exists = False
    for month in monthlyResults:
        if (companyName in month.keys()):
            exists = True
    return exists


def writeSheet(book, sheetName, banksSummary):
    resultSheet = book.create_sheet(sheetName)
    resultSheet.sheet_properties.tabColor = "ff0000"

    # СОЗДАТЬ ЗАГЛАВНУЮ СТРОКУ
    resultSheet['B2'] = 'Банк'
    headerRowFormat(resultSheet['B2'])
    resultSheet['C2'] = 'cч.51'
    headerRowFormat(resultSheet['C2'])

    # добавляем строчку по месяцам
    headerRow = 2
    monthsColumn = 4
    for monthN in e.monInt:
        cell = resultSheet.cell(row=headerRow, column=monthsColumn)
        cell.value = e.numberToMon[monthN][:3] + ". 17г."
        headerRowFormat(cell)
        monthsColumn += 1

    # заполнить значения
    dataRow = 3

    for bank_co in banksSummary:
        # полчаем название банка и номер счета
        values = bank_co.split('___')
        bankName = values[0]
        companyName = values[1]
        # Пишем счет в банке
        dataColumn = 2
        cell = resultSheet.cell(row=dataRow, column=dataColumn)
        cell.value = bankName
        # двигаемся в следующую колонку
        dataColumn += 1
        # Пишем название компании по которой делаем сумму
        cell = resultSheet.cell(row=dataRow, column=dataColumn)
        cell.value = companyName
        cell.font = cStyle.Font(bold=True)
        # двигаемся в следующую колонку
        dataColumn += 1
        # выписываем результаты
        writeMonthResults(dataRow, dataColumn, banksSummary, bank_co, resultSheet, companyName)

        # начинаем следующую строку
        dataColumn = 2
        dataRow += 1
        # опять имя банка
        cell = resultSheet.cell(row=dataRow, column=dataColumn)
        cell.value = bankName
        dataColumn += 1

        # для каждой компании в списке
        for co in e.holdingsCos:
            # убеждаемся, что компания не сама, по которой мы делаем итог и что она есть в файле
            if isNotReportingCo(co, companyName) and existsInReport(co, banksSummary[bank_co]):
                dataColumn = 3
                cell = resultSheet.cell(row=dataRow, column=dataColumn)
                cell.value = "в тч. " + co
                dataColumn += 1
                # выписываем результаты
                writeMonthResults(dataRow, dataColumn, banksSummary, bank_co, resultSheet, co)
                dataRow += 1

    # сохраняем элемент
    book.save("outputData/" + e.outpuFileName + ".xlsx")
