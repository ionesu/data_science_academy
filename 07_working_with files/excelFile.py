import openpyxl as xl
import numpy as np

file = xl.load_workbook('Sample.xlsx')
sheet = file.active

# region16Revenue = {}
#
# for row in sheet.iter_rows(min_row = 2):
#     if row[0].value:
#         region = row[0].value
#         rev2016 = row[3].value
#         defaultValue = 0
#         region16Revenue[region] = region16Revenue.get(region, defaultValue) + rev2016
#
# print(region16Revenue)

regionWPerStore = {}

for row in sheet.iter_rows(min_row=2):
    if row[0].value:
        storeName = row[0].value + " "+ row[1].value
        values = {'2015':0,'2016':0,'2017':0}

        revenue2015 = row[2].value
        revenue2016 = row[3].value
        revenue2017 = row[4].value

        values['2015'] = 1
        values['2016'] = revenue2016/revenue2015-1
        values['2017'] = revenue2017/revenue2016-1

        regionWPerStore[storeName] = values

print(regionWPerStore)


perStore = {}
perRegion = {}

#посчитаем средний рост для каждого из магазинов
for row in sheet.iter_rows(min_row=2):
    if row[0].value:
        storeName = row[1].value
        regionName = row[0].value
        values = {'2015':0,'2016':0,'2017':0}

        revenue2015 = row[2].value
        revenue2016 = row[3].value
        revenue2017 = row[4].value

        values['2015'] = 1
        values['2016'] = revenue2016/revenue2015-1
        values['2017'] = revenue2017/revenue2016-1
        values['average'] = np.mean(list(values.values()))

        perStore[storeName] = values
        perRegion[regionName] = perStore

print(perRegion["г. Москва"]["Магазин 9"])
print(perRegion["Удмуртская Республика"]["Магазин 29"]["2017"])
print(perRegion["Тюменская область"]["Магазин 18"]["average"])

#cоздадим стандартное отколнение для среднего роста по регионам, по магазинам
meanValues = []
perRegionSTDev = {}

for regionName in perRegion:
    meanValues = []

    for key, values in perRegion["г. Москва"].items():
        meanValues.append(values['average'])

    perRegionSTDev[regionName] = np.std(meanValues)#

print( perRegionSTDev['г. Москва'] )
