import openpyxl as xl
import numpy as nu

file = xl.load_workbook('Prices.xlsx')
sheet = file.active
#sheet.iter_cols(min_row=2, min_col=1)
#sheet["B"]

arraymo1Value = []

for col in sheet.iter_cols(min_row=2, min_col=2):
    for cell in col:
        arraymo1Value.append(cell.value)

print(arraymo1Value)
print(nu.mean(arraymo1Value))
print(nu.std(arraymo1Value))
